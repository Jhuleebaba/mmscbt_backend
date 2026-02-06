import os
import re
import base64
import tempfile
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO
import logging
import time
import uuid

# Document parsing libraries
import docx
from docx.document import Document
from docx.shared import Inches
import PyPDF2
import mammoth
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import html2text
from PIL import Image

logger = logging.getLogger(__name__)

class DocumentParser:
    """
    Comprehensive document parser for extracting questions from various formats
    while preserving rich text formatting for exam questions and handling instruction blocks.
    """
    
    def __init__(self):
        self.supported_formats = ['.docx', '.doc', '.pdf', '.html', '.htm', '.xlsx', '.xls']
        
    def parse_document(self, file_content: bytes, filename: str, question_type: str = 'auto') -> Dict[str, Any]:
        """
        Parse document and extract questions with their associated instructions based on type.
        
        Args:
            file_content: Raw file bytes
            filename: Original filename with extension
            question_type: 'mcq', 'theory', or 'auto' for automatic detection
            
        Returns:
            Dictionary containing parsed questions, instructions, and metadata
        """
        try:
            file_ext = os.path.splitext(filename.lower())[1]
            
            if file_ext not in self.supported_formats:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Parse based on file type
            if file_ext == '.docx':
                return self._parse_docx(file_content, question_type)
            elif file_ext == '.doc':
                return self._parse_doc_with_mammoth(file_content, question_type)
            elif file_ext == '.pdf':
                return self._parse_pdf(file_content, question_type)
            elif file_ext in ['.html', '.htm']:
                return self._parse_html(file_content, question_type)
            elif file_ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_content, question_type)
            else:
                raise ValueError(f"Parser not implemented for {file_ext}")
                
        except Exception as e:
            logger.error(f"Error parsing document {filename}: {str(e)}")
            raise
    
    def _parse_docx(self, file_content: bytes, question_type: str) -> Dict[str, Any]:
        """Parse DOCX files preserving rich text formatting."""
        temp_file_path = None
        try:
            # Create temporary file with a unique name
            temp_fd, temp_file_path = tempfile.mkstemp(suffix='.docx')
            
            # Write content to temporary file
            with os.fdopen(temp_fd, 'wb') as temp_file:
                temp_file.write(file_content)
            
            # Parse with python-docx
            doc = docx.Document(temp_file_path)
            
            # Extract content with formatting
            questions_data = self._extract_questions_from_docx(doc, question_type)
            
            # Close the document explicitly to release file handles
            del doc
            
            return questions_data
            
        except Exception as e:
            logger.error(f"Error parsing DOCX: {str(e)}")
            raise
        finally:
            # Clean up temporary file with retry mechanism for Windows
            if temp_file_path and os.path.exists(temp_file_path):
                max_retries = 5
                for attempt in range(max_retries):
                    try:
                        os.unlink(temp_file_path)
                        break
                    except (OSError, PermissionError) as e:
                        if attempt < max_retries - 1:
                            time.sleep(0.1)  # Wait 100ms before retry
                            continue
                        else:
                            logger.warning(f"Could not delete temporary file {temp_file_path}: {str(e)}")
                            # Don't raise the error, just log it as temp files will be cleaned up by OS
    
    def _parse_doc_with_mammoth(self, file_content: bytes, question_type: str) -> Dict[str, Any]:
        """Parse DOC files using mammoth for better formatting preservation."""
        try:
            # Convert to HTML with mammoth for better formatting
            result = mammoth.convert_to_html(BytesIO(file_content))
            html_content = result.value
            
            # Parse the HTML content
            return self._parse_html_content(html_content, question_type)
            
        except Exception as e:
            logger.error(f"Error parsing DOC with mammoth: {str(e)}")
            raise
    
    def _parse_pdf(self, file_content: bytes, question_type: str) -> Dict[str, Any]:
        """Parse PDF files (basic text extraction)."""
        try:
            reader = PyPDF2.PdfReader(BytesIO(file_content))
            
            full_text = ""
            for page in reader.pages:
                full_text += page.extract_text() + "\n"
            
            # Convert to basic HTML structure for consistent processing
            html_content = full_text.replace('\n', '<br>')
            
            return self._parse_html_content(html_content, question_type)
            
        except Exception as e:
            logger.error(f"Error parsing PDF: {str(e)}")
            raise
    
    def _parse_html(self, file_content: bytes, question_type: str) -> Dict[str, Any]:
        """Parse HTML files."""
        try:
            html_content = file_content.decode('utf-8')
            return self._parse_html_content(html_content, question_type)
            
        except Exception as e:
            logger.error(f"Error parsing HTML: {str(e)}")
            raise
    
    def _parse_excel(self, file_content: bytes, question_type: str) -> Dict[str, Any]:
        """Parse Excel files with structured question format."""
        try:
            workbook = load_workbook(BytesIO(file_content))
            
            mcq_questions = []
            theory_questions = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Determine sheet type based on name or content
                if 'mcq' in sheet_name.lower() or 'multiple' in sheet_name.lower():
                    mcq_questions.extend(self._parse_mcq_excel_sheet(sheet))
                elif 'theory' in sheet_name.lower() or 'essay' in sheet_name.lower():
                    theory_questions.extend(self._parse_theory_excel_sheet(sheet))
                else:
                    # Auto-detect based on content
                    auto_questions = self._parse_auto_excel_sheet(sheet, question_type)
                    mcq_questions.extend(auto_questions.get('mcq', []))
                    theory_questions.extend(auto_questions.get('theory', []))
            
            return {
                'mcq_questions': mcq_questions,
                'theory_questions': theory_questions,
                'total_questions': len(mcq_questions) + len(theory_questions),
                'format': 'excel',
                'warnings': []
            }
            
        except Exception as e:
            logger.error(f"Error parsing Excel: {str(e)}")
            raise
    
    def _extract_questions_from_docx(self, doc: Document, question_type: str) -> Dict[str, Any]:
        """Extract questions and instructions from DOCX document with rich formatting."""
        mcq_questions = []
        theory_questions = []
        instructions = []
        warnings = []
        
        current_question = None
        current_type = None
        current_instruction = None
        question_counter = 0
        
        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            
            # Convert paragraph to HTML to preserve formatting
            html_text = self._docx_paragraph_to_html(paragraph)
            
            # Debug logging
            logger.debug(f"Processing paragraph: {text[:50]}...")
            
            # First check if this is an instruction block
            instruction_match = self._detect_instruction_pattern(text)
            if instruction_match:
                logger.debug(f"Found instruction: {instruction_match['title']}")
                # Save previous question if exists
                if current_question:
                    if current_instruction:
                        current_question['instruction_id'] = current_instruction['id']
                    self._save_question(current_question, current_type, mcq_questions, theory_questions)
                    logger.debug(f"Saved previous question before instruction")
                    current_question = None
                
                # Create new instruction
                instruction_id = str(uuid.uuid4())
                current_instruction = {
                    'id': instruction_id,
                    'type': instruction_match['type'],
                    'title': instruction_match['title'],
                    'instruction_text': instruction_match.get('instruction_text', ''),
                    'full_text': html_text,
                    'applies_to': instruction_match['applies_to'],
                    'start_question': instruction_match.get('start_question'),
                    'end_question': instruction_match.get('end_question'),
                    'component': instruction_match.get('component'),
                    'identifier': instruction_match.get('identifier'),
                    'order': len(instructions)
                }
                instructions.append(current_instruction)
                continue
            
            # Check if this is a standalone instruction (not a formal instruction block)
            elif self._is_standalone_instruction(text) and not current_question:
                logger.debug(f"Found standalone instruction: {text[:30]}...")
                # Create a general instruction
                instruction_id = str(uuid.uuid4())
                current_instruction = {
                    'id': instruction_id,
                    'type': 'general',
                    'title': 'Instructions',
                    'instruction_text': text,
                    'full_text': html_text,
                    'applies_to': 'following_questions',
                    'order': len(instructions)
                }
                instructions.append(current_instruction)
                continue
            
            # Detect question patterns
            question_match = self._detect_question_pattern(text)
            
            if question_match:
                logger.debug(f"Found question pattern: {text[:50]}...")
                # Save previous question if exists
                if current_question:
                    if current_instruction:
                        current_question['instruction_id'] = current_instruction['id']
                    self._save_question(current_question, current_type, mcq_questions, theory_questions)
                    logger.debug(f"Saved previous question, count: MCQ={len(mcq_questions)}, Theory={len(theory_questions)}")
                
                # Start new question - clean the question text
                question_counter += 1
                cleaned_question_text = self._clean_question_text(text)
                current_question = {
                    'question_text': self._clean_html_preserve_formatting(html_text, is_question=True),
                    'options': [],
                    'sub_questions': [],
                    'marks': self._extract_marks(text),
                    'images': [],
                    'question_number': question_counter,
                    'instruction_id': current_instruction['id'] if current_instruction else None
                }
                current_type = self._detect_question_type(text, question_type)
                logger.debug(f"Started new question #{question_counter}, type: {current_type}")
                
            elif current_question:
                # Check if this is an option (for MCQ)
                option_match = self._detect_option_pattern(text)
                if option_match and current_type == 'mcq':
                    logger.debug(f"Found option: {text}")
                    
                    # First check if this is marked as correct answer BEFORE cleaning
                    is_correct = self._is_correct_option(text)
                    
                    # Clean the option text (remove prefixes AND correct answer markers)
                    cleaned_option_text = self._clean_html_preserve_formatting(html_text, is_option=True)
                    # Also remove correct answer markers from the cleaned text
                    cleaned_option_text = self._remove_correct_answer_markers(cleaned_option_text)
                    
                    current_question['options'].append(cleaned_option_text)
                    
                    # Set correct option index if this was marked as correct
                    if is_correct:
                        current_question['correct_option'] = len(current_question['options']) - 1
                        logger.debug(f"Marked option {len(current_question['options']) - 1} as correct, cleaned text: '{cleaned_option_text}'")
                
                # Check if this is a sub-question (for Theory)
                elif self._detect_sub_question_pattern(text) and current_type == 'theory':
                    logger.debug(f"Found sub-question: {text}")
                    sub_q = self._parse_sub_question(html_text, text)
                    current_question['sub_questions'].append(sub_q)
                
                else:
                    # Append to question text (for multi-line questions)
                    logger.debug(f"Appending to question text: {text[:30]}...")
                    additional_text = self._clean_html_preserve_formatting(html_text)
                    current_question['question_text'] += '<br>' + additional_text
            
            else:
                # This text doesn't match any pattern - could be loose text
                logger.debug(f"Unmatched text (ignoring): {text[:30]}...")
        
        # Save the last question
        if current_question:
            if current_instruction:
                current_question['instruction_id'] = current_instruction['id']
            self._save_question(current_question, current_type, mcq_questions, theory_questions)
            logger.debug(f"Saved final question")
        
        logger.info(f"Parsing complete: {len(mcq_questions)} MCQ, {len(theory_questions)} Theory, {len(instructions)} Instructions")
        
        # Process instruction ranges - assign instruction IDs based on question ranges
        self._process_instruction_ranges(instructions, mcq_questions + theory_questions)
        
        # Extract images from document
        self._extract_images_from_docx(doc, mcq_questions + theory_questions)
        
        return {
            'mcq_questions': mcq_questions,
            'theory_questions': theory_questions,
            'instructions': instructions,
            'total_questions': len(mcq_questions) + len(theory_questions),
            'total_instructions': len(instructions),
            'format': 'docx',
            'warnings': warnings
        }
    
    def _parse_html_content(self, html_content: str, question_type: str) -> Dict[str, Any]:
        """Parse HTML content to extract questions and instructions."""
        soup = BeautifulSoup(html_content, 'html.parser')
        
        mcq_questions = []
        theory_questions = []
        instructions = []
        warnings = []
        
        # Extract from HTML structure with instruction support
        result = self._extract_from_html_structure_with_instructions(soup, question_type)
        
        mcq_questions = result.get('mcq_questions', [])
        theory_questions = result.get('theory_questions', [])
        instructions = result.get('instructions', [])
        
        return {
            'mcq_questions': mcq_questions,
            'theory_questions': theory_questions,
            'instructions': instructions,
            'total_questions': len(mcq_questions) + len(theory_questions),
            'total_instructions': len(instructions),
            'format': 'html',
            'warnings': warnings
        }
    
    def _detect_question_pattern(self, text: str) -> bool:
        """Detect if text is start of a new question with very flexible patterns."""
        patterns = [
            # Standard numbered patterns
            r'^\d+[\.\)\,\:\-\s]+',  # 1. or 1) or 1, or 1: or 1- 
            r'^Q\.?\s*\d+[\.\)\,\:\-\s]*',  # Q1 or Q.1 or Q1. or Q1) or Q1:
            r'^Question\s*\.?\s*\d+[\.\)\,\:\-\s]*',  # Question 1 or Question.1 or Question 1.
            r'^\(\d+\)[\.\)\,\:\-\s]*',  # (1) or (1). or (1):
            r'^QUESTION\s*\.?\s*\d+[\.\)\,\:\-\s]*',  # QUESTION 1 or QUESTION.1
            
            # More flexible patterns
            r'^\d+[\s]*[\.\)\,\:\-][\s]*[A-Za-z]',  # 1. What is... or 1) Which of...
            r'^No\.?\s*\d+[\.\)\,\:\-\s]*',  # No.1 or No 1.
            r'^\d+[\s]*[A-Za-z]',  # 1 What is... (no punctuation)
            
            # Roman numerals
            r'^[IVX]+[\.\)\,\:\-\s]+',  # I. or II) or III:
            
            # Letter patterns for main questions (but NOT single letters that could be options)
            # Only match if it's a longer text that clearly looks like a question
        ]
        
        # Clean the text first
        text_clean = text.strip()
        
        # EXCLUDE MCQ options - these should NOT be detected as questions
        # Check if this looks like an MCQ option first
        if self._looks_like_mcq_option(text_clean):
            return False
        
        for pattern in patterns:
            if re.match(pattern, text_clean, re.IGNORECASE):
                return True
        
        # Additional check: if text starts with number/letter and is long enough to be a question
        if len(text_clean.split()) > 2:  # At least 3 words
            # Check for pattern like "1 What is..." or "A Which of the following"
            simple_patterns = [
                r'^\d+\s+[A-Z]',  # Number followed by capitalized word
                # Removed single letter pattern to avoid matching options
            ]
            for pattern in simple_patterns:
                if re.match(pattern, text_clean):
                    return True
        
        return False
    
    def _looks_like_mcq_option(self, text: str) -> bool:
        """Check if text looks like an MCQ option to avoid misclassifying as question."""
        # Short single letter options
        if re.match(r'^[A-Da-d][\.\)\,\:\-\s]+', text):
            # If it's A., B., C., D. followed by a short text, it's likely an option
            words = text.split()
            if len(words) <= 5:  # Short text, likely an option
                return True
            # If it contains option-like words
            option_words = ['option', 'choice', 'answer', 'correct', 'wrong', 'false', 'true']
            if any(word.lower() in text.lower() for word in option_words):
                return True
        
        # Numbered options (1., 2., 3., 4.)
        if re.match(r'^[1-4][\.\)\,\:\-\s]+', text):
            words = text.split()
            if len(words) <= 5:  # Short numbered option
                return True
        
        return False
    
    def _detect_question_type(self, text: str, default_type: str) -> str:
        """Detect question type from text content."""
        if default_type in ['mcq', 'theory']:
            return default_type
        
        # Auto-detection patterns
        mcq_indicators = ['select', 'choose', 'pick', 'option', 'a)', 'b)', 'c)', 'd)', 'A)', 'B)', 'C)', 'D)']
        theory_indicators = ['explain', 'describe', 'discuss', 'analyze', 'elaborate', 'write']
        
        text_lower = text.lower()
        
        # Check for MCQ indicators
        mcq_score = sum(1 for indicator in mcq_indicators if indicator in text_lower)
        theory_score = sum(1 for indicator in theory_indicators if indicator in text_lower)
        
        # Additional heuristics for MCQ detection
        # If question has blanks/fill-ins and no explicit theory indicators, likely MCQ
        if ('_____' in text or '__' in text) and theory_score == 0:
            mcq_score += 2
        
        # If question is relatively short (under 20 words) and has blanks, likely MCQ
        if len(text.split()) < 20 and ('_____' in text or '__' in text):
            mcq_score += 1
        
        # If question asks for completion/filling, likely MCQ
        completion_indicators = ['complete', 'fill', 'blank', 'missing']
        if any(indicator in text_lower for indicator in completion_indicators):
            mcq_score += 1
        
        detected_type = 'mcq' if mcq_score > theory_score else 'theory'
        logger.debug(f"Question type detection: '{text[:50]}...' -> {detected_type} (MCQ:{mcq_score}, Theory:{theory_score})")
        
        return detected_type
    
    def _detect_option_pattern(self, text: str) -> bool:
        """Detect if text is an MCQ option with very flexible patterns."""
        patterns = [
            # Standard option patterns with various punctuation
            r'^[a-d][\.\)\,\:\-\s]+',  # a. or a) or a, or a: or a-
            r'^[A-D][\.\)\,\:\-\s]+',  # A. or A) or A, or A: or A-
            r'^\([a-dA-D]\)[\.\,\:\-\s]*',  # (a) or (A) with optional punctuation after
            r'^[a-dA-D][\.\,\:\-\s]*',  # a) or A) with optional punctuation
            
            # More flexible patterns
            r'^[a-dA-D][\s]*[\.\)\,\:\-][\s]*[A-Za-z]',  # Letter + punctuation + word
            r'^[a-dA-D][\s]+[A-Za-z]',  # Letter + space + word (no punctuation)
            
            # Numbered options (1, 2, 3, 4)
            r'^[1-4][\.\)\,\:\-\s]+',  # 1. or 2) or 3, etc.
            r'^\([1-4]\)[\.\,\:\-\s]*',  # (1) or (2) etc.
            
            # Roman numerals for options
            r'^[ivx]+[\.\)\,\:\-\s]+',  # i. or ii) or iii:
            
            # Bullet points or dashes
            r'^[\-\•\*][\s]*[A-Za-z]',  # - Option or • Option or * Option
        ]
        
        text_clean = text.strip()
        
        for pattern in patterns:
            if re.match(pattern, text_clean, re.IGNORECASE):
                return True
        
        # Additional check for options that start with common option indicators
        option_starters = ['a', 'b', 'c', 'd', 'A', 'B', 'C', 'D', '1', '2', '3', '4']
        first_char = text_clean[0] if text_clean else ''
        
        if first_char in option_starters and len(text_clean.split()) > 1:
            # Check if the second character is punctuation or space
            if len(text_clean) > 1:
                second_char = text_clean[1]
                if second_char in '.,):- ' or text_clean[1:3] == ') ':
                    return True
        
        return False
    
    def _detect_sub_question_pattern(self, text: str) -> bool:
        """Detect if text is a theory sub-question."""
        patterns = [
            r'^[a-z][\.\)]\s+',  # a. or a)
            r'^\([a-z]\)\s+',  # (a)
            r'^[ivx]+[\.\)]\s+',  # i. ii. iii.
            r'^\([ivx]+\)\s+',  # (i) (ii)
        ]
        
        for pattern in patterns:
            if re.match(pattern, text):
                return True
        return False
    
    def _is_correct_option(self, text: str) -> bool:
        """Check if option is marked as correct with very flexible patterns."""
        indicators = [
            '*', '✓', '√', '✗', '×', '▪', '■', '◾', '⬛',  # Symbols
            'correct', 'answer', '(correct)', '[correct]', 
            '(answer)', '[answer]', '(right)', '[right]',
            'right', 'true', '(true)', '[true]',
            '(✓)', '[✓]', '(*)', '[*]',  # Bracketed symbols
            'ans', '(ans)', '[ans]',  # Short forms
        ]
        
        text_lower = text.lower().strip()
        
        # Check for indicators at the end of the text (most common)
        for indicator in indicators:
            if text_lower.endswith(indicator.lower()):
                return True
            # Also check if indicator is within parentheses or brackets at the end
            if text_lower.endswith(f'({indicator.lower()})') or text_lower.endswith(f'[{indicator.lower()}]'):
                return True
        
        # Check for indicators anywhere in the text
        for indicator in indicators:
            if indicator.lower() in text_lower:
                return True
        
        # Check for bold or emphasized text patterns (in HTML context)
        if '<strong>' in text or '<b>' in text or '<em>' in text:
            return True
        
        # Check for all caps option (might indicate correct answer)
        words = text_lower.split()
        if len(words) > 1:
            # If more than half the words are in caps in original text, might be correct
            caps_count = sum(1 for word in text.split() if word.isupper())
            if caps_count > len(words) / 2:
                return True
        
        return False
    
    def _extract_marks(self, text: str) -> int:
        """Extract marks from question text with very flexible patterns."""
        # Look for patterns like [5 marks], (10 points), 5pts, etc.
        patterns = [
            # Standard patterns
            r'\[(\d+)\s*marks?\]',
            r'\((\d+)\s*marks?\)',
            r'\[(\d+)\s*points?\]',
            r'\((\d+)\s*points?\)',
            r'\[(\d+)\s*pts?\]',
            r'\((\d+)\s*pts?\)',
            
            # More flexible patterns
            r'(\d+)\s*marks?',
            r'(\d+)\s*points?',
            r'(\d+)\s*pts?',
            r'(\d+)\s*mark',
            r'(\d+)\s*point',
            r'(\d+)\s*pt',
            
            # With separators
            r'[-–—]\s*(\d+)\s*marks?',
            r'[-–—]\s*(\d+)\s*points?',
            r'[-–—]\s*(\d+)\s*pts?',
            
            # At the end with various punctuation
            r'.*[^\d](\d+)m\s*$',  # 5m at the end
            r'.*[^\d](\d+)p\s*$',  # 5p at the end
            
            # With colons or equals
            r'marks?\s*[:=]\s*(\d+)',
            r'points?\s*[:=]\s*(\d+)',
            r'total\s*[:=]\s*(\d+)',
        ]
        
        text_lower = text.lower()
        
        for pattern in patterns:
            match = re.search(pattern, text_lower, re.IGNORECASE)
            if match:
                try:
                    marks = int(match.group(1))
                    if 1 <= marks <= 100:  # Reasonable range for marks
                        return marks
                except (ValueError, IndexError):
                    continue
        
        # If no marks found, try to extract any number that might represent marks
        # Look for standalone numbers in reasonable range
        numbers = re.findall(r'\b(\d+)\b', text)
        for num_str in numbers:
            try:
                num = int(num_str)
                if 1 <= num <= 20:  # Most common range for individual question marks
                    return num
            except ValueError:
                continue
        
        return 1  # Default marks
    
    def _parse_sub_question(self, html_text: str, raw_text: str) -> Dict[str, Any]:
        """Parse a theory sub-question."""
        # Extract sub-question number
        sub_number_match = re.match(r'^([a-z]+|[ivx]+)[\.\)]\s*', raw_text)
        sub_number = sub_number_match.group(1) if sub_number_match else 'a'
        
        # Clean the sub-question text by removing the prefix
        cleaned_text = self._clean_option_text(raw_text)  # Same logic as options
        
        # Preserve any formatting from HTML
        if html_text.strip() != raw_text.strip():
            # Apply formatting to cleaned text if original had formatting
            soup = BeautifulSoup(html_text, 'html.parser')
            has_bold = '<strong>' in html_text or '<b>' in html_text
            has_italic = '<em>' in html_text or '<i>' in html_text
            has_underline = '<u>' in html_text
            
            if has_bold:
                cleaned_text = f'<strong>{cleaned_text}</strong>'
            if has_italic:
                cleaned_text = f'<em>{cleaned_text}</em>'
            if has_underline:
                cleaned_text = f'<u>{cleaned_text}</u>'
        
        return {
            'sub_number': sub_number,
            'sub_text': cleaned_text,
            'sub_marks': self._extract_marks(raw_text) or 1
        }
    
    def _save_question(self, question: Dict, q_type: str, mcq_list: List, theory_list: List):
        """Save question to appropriate list with instruction metadata."""
        logger.debug(f"Attempting to save question type: {q_type}")
        logger.debug(f"Question data: text='{question['question_text'][:50]}...', options={len(question.get('options', []))}")
        
        if q_type == 'mcq':
            # Validate MCQ question
            if len(question['options']) < 2:
                logger.warning(f"Skipping MCQ with insufficient options: {len(question['options'])}")
                return  # Skip invalid MCQ
            
            if 'correct_option' not in question:
                logger.debug("No correct option specified, defaulting to 0")
                question['correct_option'] = 0  # Default to first option
            
            mcq_question = {
                'question_text': question['question_text'],
                'question_type': 'mcq',
                'options': question['options'],
                'correct_option': question['correct_option'],
                'marks': question['marks'],
                'instruction_id': question.get('instruction_id'),
                'question_number': question.get('question_number')
            }
            mcq_list.append(mcq_question)
            logger.info(f"Saved MCQ question #{question.get('question_number')} with {len(question['options'])} options")
        
        else:  # theory
            # If no sub-questions, create a default one
            if not question['sub_questions']:
                logger.debug("Creating default sub-question for theory question")
                question['sub_questions'] = [{
                    'sub_number': 'a',
                    'sub_text': question['question_text'],
                    'sub_marks': question['marks']
                }]
                question['question_text'] = ''  # Main question text is now in sub-question
            
            theory_question = {
                'question_text': question['question_text'],
                'question_type': 'theory',
                'sub_questions': question['sub_questions'],
                'marks': sum(sq['sub_marks'] for sq in question['sub_questions']),
                'instruction_id': question.get('instruction_id'),
                'question_number': question.get('question_number')
            }
            theory_list.append(theory_question)
            logger.info(f"Saved Theory question #{question.get('question_number')} with {len(question['sub_questions'])} sub-questions")
    
    def _docx_paragraph_to_html(self, paragraph) -> str:
        """Convert DOCX paragraph to HTML preserving basic formatting."""
        html = ""
        
        for run in paragraph.runs:
            text = run.text
            if not text:
                continue
            
            # Apply formatting
            if run.bold:
                text = f"<strong>{text}</strong>"
            if run.italic:
                text = f"<em>{text}</em>"
            if run.underline:
                text = f"<u>{text}</u>"
            
            html += text
        
        return html
    
    def _extract_images_from_docx(self, doc: Document, questions: List[Dict]):
        """Extract images from DOCX and embed as base64."""
        try:
            # This is a simplified version - full implementation would need
            # to map images to specific questions based on position
            pass
        except Exception as e:
            logger.warning(f"Could not extract images: {str(e)}")
    
    def _parse_mcq_excel_sheet(self, sheet) -> List[Dict]:
        """Parse Excel sheet formatted for MCQ questions."""
        questions = []
        
        # Expected columns: Question, Option A, Option B, Option C, Option D, Correct, Marks
        header_row = 1
        
        for row in range(header_row + 1, sheet.max_row + 1):
            question_text = sheet.cell(row, 1).value
            if not question_text:
                continue
            
            options = []
            option_correct_flags = []  # Track which options are marked as correct
            for col in range(2, 6):  # Columns B-E for options
                option = sheet.cell(row, col).value
                if option:
                    option_text = str(option)
                    # Check if this option is marked as correct
                    is_correct = self._is_correct_option(option_text)
                    option_correct_flags.append(is_correct)
                    
                    # Clean the option text (remove correct answer markers)
                    cleaned_option = self._remove_correct_answer_markers(option_text)
                    options.append(cleaned_option)
                else:
                    option_correct_flags.append(False)
            
            if len(options) < 2:
                continue
            
            # Get correct answer - check both the "Correct" column and in-text markers
            correct_option = 0
            
            # First check if any option was marked as correct in the text
            if any(option_correct_flags):
                correct_option = next(i for i, is_correct in enumerate(option_correct_flags) if is_correct)
            else:
                # Fall back to the "Correct" column
                correct_answer = sheet.cell(row, 6).value
                if correct_answer:
                    correct_answer = str(correct_answer).upper()
                    option_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
                    correct_option = option_map.get(correct_answer, 0)
            
            # Get marks
            marks = sheet.cell(row, 7).value or 1
            
            questions.append({
                'question_text': str(question_text),
                'question_type': 'mcq',
                'options': options,
                'correct_option': correct_option,
                'marks': int(marks) if isinstance(marks, (int, float)) else 1
            })
        
        return questions
    
    def _parse_theory_excel_sheet(self, sheet) -> List[Dict]:
        """Parse Excel sheet formatted for Theory questions."""
        questions = []
        
        # Expected columns: Question, Sub-questions (JSON or separated), Marks
        header_row = 1
        
        for row in range(header_row + 1, sheet.max_row + 1):
            question_text = sheet.cell(row, 1).value
            if not question_text:
                continue
            
            # Parse sub-questions if provided
            sub_questions_data = sheet.cell(row, 2).value
            sub_questions = []
            
            if sub_questions_data:
                # Try to parse as structured data
                try:
                    import json
                    sub_questions = json.loads(sub_questions_data)
                except:
                    # Parse as simple text with patterns
                    sub_questions = self._parse_sub_questions_text(str(sub_questions_data))
            
            if not sub_questions:
                # Create default sub-question
                marks = sheet.cell(row, 3).value or 1
                sub_questions = [{
                    'sub_number': 'a',
                    'sub_text': str(question_text),
                    'sub_marks': int(marks) if isinstance(marks, (int, float)) else 1
                }]
                question_text = ''
            
            questions.append({
                'question_text': str(question_text) if question_text else '',
                'question_type': 'theory',
                'sub_questions': sub_questions,
                'marks': sum(sq.get('sub_marks', 1) for sq in sub_questions)
            })
        
        return questions
    
    def _parse_auto_excel_sheet(self, sheet, question_type: str) -> Dict[str, List]:
        """Auto-detect and parse Excel sheet questions."""
        # This would implement auto-detection logic
        # For now, return empty
        return {'mcq': [], 'theory': []}
    
    def _parse_sub_questions_text(self, text: str) -> List[Dict]:
        """Parse sub-questions from text."""
        sub_questions = []
        
        # Split by common patterns
        parts = re.split(r'[a-z][\.\)]\s*', text)
        
        for i, part in enumerate(parts[1:], 1):  # Skip empty first part
            if part.strip():
                sub_questions.append({
                    'sub_number': chr(96 + i),  # a, b, c, etc.
                    'sub_text': part.strip(),
                    'sub_marks': 1
                })
        
        return sub_questions
    
    def _extract_from_html_structure_with_instructions(self, soup: BeautifulSoup, question_type: str) -> Dict[str, List]:
        """Extract questions and instructions from HTML structure."""
        mcq_questions = []
        theory_questions = []
        instructions = []
        
        # Look for common HTML patterns
        elements = soup.find_all(['p', 'div', 'li', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        
        current_question = None
        current_instruction = None
        question_counter = 0
        
        for element in elements:
            text = element.get_text().strip()
            if not text:
                continue
            
            # Check for instruction patterns first
            instruction_match = self._detect_instruction_pattern(text)
            if instruction_match:
                # Save current question if exists
                if current_question:
                    if current_instruction:
                        current_question['instruction_id'] = current_instruction['id']
                    self._save_html_question(current_question, mcq_questions, theory_questions)
                    current_question = None
                
                # Create instruction
                instruction_id = str(uuid.uuid4())
                current_instruction = {
                    'id': instruction_id,
                    'type': instruction_match['type'],
                    'title': instruction_match['title'],
                    'instruction_text': instruction_match.get('instruction_text', ''),
                    'full_text': str(element),
                    'applies_to': instruction_match['applies_to'],
                    'start_question': instruction_match.get('start_question'),
                    'end_question': instruction_match.get('end_question'),
                    'component': instruction_match.get('component'),
                    'identifier': instruction_match.get('identifier'),
                    'order': len(instructions)
                }
                instructions.append(current_instruction)
                continue
            
            # Check for standalone instructions
            elif self._is_standalone_instruction(text) and not current_question:
                instruction_id = str(uuid.uuid4())
                current_instruction = {
                    'id': instruction_id,
                    'type': 'general',
                    'title': 'Instructions',
                    'instruction_text': text,
                    'full_text': str(element),
                    'applies_to': 'following_questions',
                    'order': len(instructions)
                }
                instructions.append(current_instruction)
                continue
            
            # Check for question patterns
            if self._detect_question_pattern(text):
                if current_question:
                    if current_instruction:
                        current_question['instruction_id'] = current_instruction['id']
                    self._save_html_question(current_question, mcq_questions, theory_questions)
                
                question_counter += 1
                current_question = {
                    'question_text': str(element),
                    'type': self._detect_question_type(text, question_type),
                    'options': [],
                    'sub_questions': [],
                    'marks': self._extract_marks(text),
                    'question_number': question_counter,
                    'instruction_id': current_instruction['id'] if current_instruction else None
                }
            
            elif current_question:
                if self._detect_option_pattern(text) and current_question['type'] == 'mcq':
                    # Check if this is marked as correct answer BEFORE cleaning
                    is_correct = self._is_correct_option(text)
                    
                    # Clean the option text and remove correct answer markers
                    cleaned_option_text = self._clean_html_preserve_formatting(str(element), is_option=True)
                    cleaned_option_text = self._remove_correct_answer_markers(cleaned_option_text)
                    
                    current_question['options'].append(cleaned_option_text)
                    
                    if is_correct:
                        current_question['correct_option'] = len(current_question['options']) - 1
        
        if current_question:
            if current_instruction:
                current_question['instruction_id'] = current_instruction['id']
            self._save_html_question(current_question, mcq_questions, theory_questions)
        
        # Process instruction ranges
        self._process_instruction_ranges(instructions, mcq_questions + theory_questions)
        
        return {
            'mcq_questions': mcq_questions,
            'theory_questions': theory_questions,
            'instructions': instructions
        }
    
    def _save_html_question(self, question: Dict, mcq_list: List, theory_list: List):
        """Save HTML question to appropriate list."""
        if question['type'] == 'mcq':
            if len(question['options']) < 2:
                return  # Skip invalid MCQ
            
            mcq_list.append({
                'question_text': self._clean_html_preserve_formatting(question['question_text'], is_question=True),
                'question_type': 'mcq',
                'options': [self._clean_html_preserve_formatting(opt, is_option=True) for opt in question['options']],
                'correct_option': question.get('correct_option', 0),
                'marks': question['marks'],
                'instruction_id': question.get('instruction_id'),
                'question_number': question.get('question_number')
            })
        else:
            # Theory question
            sub_questions = question['sub_questions'] or [{
                'sub_number': 'a',
                'sub_text': self._clean_html_preserve_formatting(question['question_text'], is_question=True),
                'sub_marks': question['marks']
            }]
            
            theory_list.append({
                'question_text': self._clean_html_preserve_formatting(question['question_text'], is_question=True),
                'question_type': 'theory',
                'sub_questions': sub_questions,
                'marks': question['marks'],
                'instruction_id': question.get('instruction_id'),
                'question_number': question.get('question_number')
            })
    
    def _extract_from_html_structure(self, soup: BeautifulSoup, question_type: str) -> List[Dict]:
        """Extract questions from HTML structure (legacy method)."""
        result = self._extract_from_html_structure_with_instructions(soup, question_type)
        
        # Convert to legacy format for backward compatibility
        questions = []
        for q in result['mcq_questions']:
            questions.append({
                'question_text': q['question_text'],
                'question_type': 'mcq',
                'options': q['options'],
                'correct_option': q['correct_option'],
                'marks': q['marks'],
                'type': 'mcq'
            })
        
        for q in result['theory_questions']:
            questions.append({
                'question_text': q['question_text'],
                'question_type': 'theory',
                'sub_questions': q['sub_questions'],
                'marks': q['marks'],
                'type': 'theory'
            })
        
        return questions
    
    def _clean_html(self, html_text: str) -> str:
        """Clean and preserve essential HTML formatting."""
        soup = BeautifulSoup(html_text, 'html.parser')
        
        # Remove unwanted attributes but keep basic formatting
        for tag in soup.find_all():
            # Keep only essential attributes
            allowed_attrs = ['href', 'src', 'alt', 'title']
            tag.attrs = {k: v for k, v in tag.attrs.items() if k in allowed_attrs}
        
        return str(soup)

    def _detect_instruction_pattern(self, text: str) -> Optional[Dict[str, Any]]:
        """
        Detect if text is an instruction block.
        
        Patterns recognized:
        - INSTRUCTIONS: or Instruction:
        - SECTION A: SYNONYMS or Section A - Grammar
        - Component 1: Antonyms
        - Instructions for Questions 1-10:
        - COMPREHENSION: Read the passage and answer...
        - VOCABULARY: Choose the correct meaning...
        """
        instruction_patterns = [
            # General instruction patterns
            r'^INSTRUCTIONS?\s*:?\s*(.+)',
            r'^INSTRUCTION\s+FOR\s+(.+)',
            r'^GENERAL\s+INSTRUCTIONS?\s*:?\s*(.+)',
            
            # Section-based patterns  
            r'^SECTION\s+([A-Z])\s*[-:]?\s*(.+)',
            r'^PART\s+([A-Z]|[IVX]+|\d+)\s*[-:]?\s*(.+)',
            r'^COMPONENT\s+(\d+)\s*[-:]?\s*(.+)',
            
            # Subject-specific patterns
            r'^(SYNONYMS?|ANTONYMS?|GRAMMAR|COMPREHENSION|VOCABULARY|LEXIS|STRUCTURE)\s*:?\s*(.+)',
            r'^(READING|WRITING|LISTENING|SPEAKING)\s+SECTION\s*:?\s*(.+)',
            
            # Question range patterns
            r'^INSTRUCTIONS?\s+FOR\s+QUESTIONS?\s+(\d+)\s*[-–]\s*(\d+)\s*:?\s*(.+)',
            r'^FOR\s+QUESTIONS?\s+(\d+)\s*[-–]\s*(\d+)\s*:?\s*(.+)',
            
            # Component with description
            r'^([A-Z][A-Z\s&]+):\s*(.+)',  # LEXIS AND STRUCTURE: Choose the correct...
        ]
        
        text_stripped = text.strip()
        
        for pattern in instruction_patterns:
            match = re.match(pattern, text_stripped, re.IGNORECASE | re.MULTILINE)
            if match:
                groups = match.groups()
                
                # Determine instruction type and metadata
                if 'SECTION' in pattern.upper():
                    return {
                        'type': 'section',
                        'identifier': groups[0] if len(groups) > 1 else 'A',
                        'title': groups[1] if len(groups) > 1 else groups[0],
                        'full_text': text_stripped,
                        'applies_to': 'following_questions'
                    }
                elif 'COMPONENT' in pattern.upper() or 'PART' in pattern.upper():
                    return {
                        'type': 'component',
                        'identifier': groups[0] if len(groups) > 1 else '1',
                        'title': groups[1] if len(groups) > 1 else groups[0],
                        'full_text': text_stripped,
                        'applies_to': 'following_questions'
                    }
                elif 'QUESTIONS' in pattern.upper() and len(groups) >= 3:
                    return {
                        'type': 'range',
                        'start_question': int(groups[0]),
                        'end_question': int(groups[1]),
                        'title': f"Questions {groups[0]}-{groups[1]}",
                        'instruction_text': groups[2],
                        'full_text': text_stripped,
                        'applies_to': 'question_range'
                    }
                elif any(keyword in pattern.upper() for keyword in ['SYNONYMS', 'ANTONYMS', 'GRAMMAR', 'VOCABULARY']):
                    subject_component = groups[0] if groups else 'General'
                    instruction_text = groups[1] if len(groups) > 1 else text_stripped
                    return {
                        'type': 'subject_component',
                        'component': subject_component.title(),
                        'title': subject_component.title(),
                        'instruction_text': instruction_text,
                        'full_text': text_stripped,
                        'applies_to': 'following_questions'
                    }
                else:
                    return {
                        'type': 'general',
                        'title': groups[0] if groups else 'Instructions',
                        'instruction_text': groups[1] if len(groups) > 1 else groups[0] if groups else text_stripped,
                        'full_text': text_stripped,
                        'applies_to': 'following_questions'
                    }
        
        return None
    
    def _is_standalone_instruction(self, text: str) -> bool:
        """Check if text is a standalone instruction (not part of a question)."""
        # Instructions are usually:
        # 1. In all caps or title case
        # 2. Don't start with question numbers
        # 3. Contain instruction keywords
        # 4. Are generally descriptive rather than interrogative
        
        text_stripped = text.strip()
        
        # Skip if it starts with question patterns
        if self._detect_question_pattern(text_stripped):
            return False
        
        # Check for instruction keywords
        instruction_keywords = [
            'choose', 'select', 'identify', 'complete', 'fill', 'match',
            'read the passage', 'answer the following', 'from the options',
            'correct answer', 'best answer', 'most appropriate',
            'instructions', 'direction', 'note', 'read carefully'
        ]
        
        text_lower = text_stripped.lower()
        has_instruction_keywords = any(keyword in text_lower for keyword in instruction_keywords)
        
        # Check if it's likely an instruction based on structure
        is_descriptive = (
            text_stripped.endswith(':') or 
            text_stripped.endswith('.') or
            'choose' in text_lower or
            'select' in text_lower or
            'following' in text_lower
        )
        
        # Check length - instructions are usually longer explanatory text
        is_substantial = len(text_stripped.split()) > 3
        
        return has_instruction_keywords and is_descriptive and is_substantial
    
    def _process_instruction_ranges(self, instructions: List[Dict], all_questions: List[Dict]):
        """Process instruction ranges and assign instruction IDs to questions based on ranges."""
        for instruction in instructions:
            if instruction['applies_to'] == 'question_range':
                start_q = instruction.get('start_question', 1)
                end_q = instruction.get('end_question', len(all_questions))
                
                # Assign this instruction to questions in the range
                for question in all_questions:
                    q_num = question.get('question_number', 0)
                    if start_q <= q_num <= end_q:
                        # Only assign if the question doesn't already have an instruction
                        if not question.get('instruction_id'):
                            question['instruction_id'] = instruction['id']

    def _clean_question_text(self, text: str) -> str:
        """Remove question number/prefix from question text."""
        # Patterns to remove from the beginning of question text
        patterns_to_remove = [
            r'^\d+[\.\)\,\:\-\s]+',  # 1. or 1) or 1, or 1: or 1-
            r'^Q\.?\s*\d+[\.\)\,\:\-\s]*',  # Q1 or Q.1 or Q1. or Q1)
            r'^Question\s*\.?\s*\d+[\.\)\,\:\-\s]*',  # Question 1 or Question.1
            r'^\(\d+\)[\.\)\,\:\-\s]*',  # (1) or (1).
            r'^QUESTION\s*\.?\s*\d+[\.\)\,\:\-\s]*',  # QUESTION 1
            r'^No\.?\s*\d+[\.\)\,\:\-\s]*',  # No.1 or No 1.
            r'^[IVX]+[\.\)\,\:\-\s]+',  # I. or II) or III:
        ]
        
        cleaned_text = text.strip()
        
        for pattern in patterns_to_remove:
            cleaned_text = re.sub(pattern, '', cleaned_text, flags=re.IGNORECASE).strip()
        
        return cleaned_text
    
    def _clean_option_text(self, text: str) -> str:
        """Remove option letter/number prefix from option text."""
        # Patterns to remove from the beginning of option text
        patterns_to_remove = [
            r'^[a-dA-D][\.\)\,\:\-\s]+',  # a. or A) or a, or A:
            r'^\([a-dA-D]\)[\.\,\:\-\s]*',  # (a) or (A) with optional punctuation
            r'^[1-4][\.\)\,\:\-\s]+',  # 1. or 2) or 3,
            r'^\([1-4]\)[\.\,\:\-\s]*',  # (1) or (2)
            r'^[ivx]+[\.\)\,\:\-\s]+',  # i. or ii) or iii:
            r'^[\-\•\*][\s]*',  # - or • or * bullet points
        ]
        
        cleaned_text = text.strip()
        
        for pattern in patterns_to_remove:
            cleaned_text = re.sub(pattern, '', cleaned_text, flags=re.IGNORECASE).strip()
        
        return cleaned_text
    
    def _remove_correct_answer_markers(self, text: str) -> str:
        """Remove correct answer markers from option text."""
        # List of indicators to remove (same as in _is_correct_option)
        indicators_to_remove = [
            r'\*',  # Asterisk
            r'✓', r'√', r'✗', r'×', r'▪', r'■', r'◾', r'⬛',  # Symbols
            r'\(correct\)', r'\[correct\]', r'correct',  # Correct variations
            r'\(answer\)', r'\[answer\]', r'answer',  # Answer variations
            r'\(right\)', r'\[right\]', r'right',  # Right variations
            r'\(true\)', r'\[true\]', r'true',  # True variations
            r'\(✓\)', r'\[✓\]', r'\(\*\)', r'\[\*\]',  # Bracketed symbols
            r'\(ans\)', r'\[ans\]', r'ans',  # Short forms
        ]
        
        cleaned_text = text.strip()
        
        # Remove indicators from anywhere in the text (beginning, middle, or end)
        for indicator in indicators_to_remove:
            # Remove the indicator with optional surrounding spaces
            pattern = r'\s*' + indicator + r'\s*'
            cleaned_text = re.sub(pattern, ' ', cleaned_text, flags=re.IGNORECASE)
        
        # Clean up extra spaces and normalize whitespace
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        
        # Remove trailing punctuation that might be left after removing markers
        cleaned_text = re.sub(r'\s*[,;]\s*$', '', cleaned_text)
        
        return cleaned_text

    def _clean_html_preserve_formatting(self, html_text: str, is_question: bool = False, is_option: bool = False) -> str:
        """Clean HTML while preserving formatting, and remove prefixes for questions or options."""
        # First, extract the plain text to clean prefixes
        soup = BeautifulSoup(html_text, 'html.parser')
        plain_text = soup.get_text().strip()
        
        # Clean the text based on type
        if is_question:
            cleaned_text = self._clean_question_text(plain_text)
            # Remove marks notation from question text
            cleaned_text = self._remove_marks_from_text(cleaned_text)
        elif is_option:
            cleaned_text = self._clean_option_text(plain_text)
        else:
            cleaned_text = plain_text
        
        # Only preserve formatting if the original text actually had meaningful formatting
        # and the cleaned text is substantially the same as the original
        if html_text.strip() != plain_text and len(cleaned_text.strip()) > 0:
            # Check if formatting should be preserved (avoid adding random formatting)
            original_formatted_text = soup.get_text()
            
            # Only apply formatting if it seems intentional (e.g., specific words are formatted)
            # Look for formatting tags that are around specific content, not the whole text
            has_selective_bold = bool(re.search(r'<(strong|b)>[^<]{1,50}</(strong|b)>', html_text))
            has_selective_italic = bool(re.search(r'<(em|i)>[^<]{1,50}</(em|i)>', html_text))
            
            # Preserve bold, italic, and underline formatting
            has_selective_underline = bool(re.search(r'<u>[^<]{1,50}</u>', html_text))
            
            if has_selective_bold and '<strong>' in html_text:
                # Try to preserve bold formatting on specific parts
                bold_content = re.findall(r'<strong>([^<]+)</strong>', html_text)
                for content in bold_content:
                    if content.strip() in cleaned_text:
                        cleaned_text = cleaned_text.replace(content.strip(), f'<strong>{content.strip()}</strong>')
            
            if has_selective_italic and '<em>' in html_text:
                # Try to preserve italic formatting on specific parts
                italic_content = re.findall(r'<em>([^<]+)</em>', html_text)
                for content in italic_content:
                    if content.strip() in cleaned_text:
                        cleaned_text = cleaned_text.replace(content.strip(), f'<em>{content.strip()}</em>')
            
            if has_selective_underline and '<u>' in html_text:
                # Try to preserve underline formatting on specific parts
                underline_content = re.findall(r'<u>([^<]+)</u>', html_text)
                for content in underline_content:
                    if content.strip() in cleaned_text:
                        cleaned_text = cleaned_text.replace(content.strip(), f'<u>{content.strip()}</u>')
        
        return cleaned_text
    
    def _remove_marks_from_text(self, text: str) -> str:
        """Remove marks notation from text after extraction."""
        # Remove patterns like [5 marks], (10 points), 5pts, etc.
        patterns_to_remove = [
            r'\[(\d+)\s*marks?\]',
            r'\((\d+)\s*marks?\)',
            r'\[(\d+)\s*points?\]',
            r'\((\d+)\s*points?\)',
            r'\[(\d+)\s*pts?\]',
            r'\((\d+)\s*pts?\)',
            r'[-–—]\s*(\d+)\s*marks?',
            r'[-–—]\s*(\d+)\s*points?',
            r'[-–—]\s*(\d+)\s*pts?',
            r'\b(\d+)\s*marks?\b',
            r'\b(\d+)\s*points?\b',
            r'\b(\d+)\s*pts?\b',
            r'\b(\d+)m\b',
            r'\b(\d+)p\b',
            r'marks?\s*[:=]\s*(\d+)',
            r'points?\s*[:=]\s*(\d+)',
            r'total\s*[:=]\s*(\d+)',
        ]
        
        cleaned_text = text
        for pattern in patterns_to_remove:
            cleaned_text = re.sub(pattern, '', cleaned_text, flags=re.IGNORECASE)
        
        # Clean up extra spaces and punctuation
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        cleaned_text = re.sub(r'\s*[,;]\s*$', '', cleaned_text)  # Remove trailing punctuation
        
        return cleaned_text


class QuestionValidator:
    """Validate parsed questions before database insertion."""
    
    @staticmethod
    def validate_mcq_question(question: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate MCQ question structure with instruction support."""
        errors = []
        
        if not question.get('question_text', '').strip():
            errors.append("Question text is required")
        
        options = question.get('options', [])
        if len(options) < 2:
            errors.append("MCQ must have at least 2 options")
        
        correct_option = question.get('correct_option')
        if correct_option is None or correct_option < 0 or correct_option >= len(options):
            errors.append("Valid correct option must be specified")
        
        marks = question.get('marks', 0)
        if marks <= 0:
            errors.append("Marks must be positive")
        
        # Instruction ID is optional but should be string if provided
        instruction_id = question.get('instruction_id')
        if instruction_id is not None and not isinstance(instruction_id, str):
            errors.append("Instruction ID must be a string")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def validate_theory_question(question: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate Theory question structure with instruction support."""
        errors = []
        
        sub_questions = question.get('sub_questions', [])
        if not sub_questions:
            errors.append("Theory question must have at least one sub-question")
        
        total_marks = 0
        for i, sub_q in enumerate(sub_questions):
            if not sub_q.get('sub_text', '').strip():
                errors.append(f"Sub-question {i+1} text is required")
            
            sub_marks = sub_q.get('sub_marks', 0)
            if sub_marks <= 0:
                errors.append(f"Sub-question {i+1} marks must be positive")
            
            total_marks += sub_marks
        
        if total_marks <= 0:
            errors.append("Total marks must be positive")
        
        # Instruction ID is optional but should be string if provided
        instruction_id = question.get('instruction_id')
        if instruction_id is not None and not isinstance(instruction_id, str):
            errors.append("Instruction ID must be a string")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def validate_instruction(instruction: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate instruction structure."""
        errors = []
        
        if not instruction.get('id'):
            errors.append("Instruction ID is required")
        
        if not instruction.get('title', '').strip():
            errors.append("Instruction title is required")
        
        if not instruction.get('type'):
            errors.append("Instruction type is required")
        
        valid_types = ['general', 'section', 'component', 'range', 'subject_component']
        if instruction.get('type') not in valid_types:
            errors.append(f"Instruction type must be one of: {', '.join(valid_types)}")
        
        applies_to = instruction.get('applies_to')
        valid_applies_to = ['following_questions', 'question_range', 'all_questions']
        if applies_to not in valid_applies_to:
            errors.append(f"Applies_to must be one of: {', '.join(valid_applies_to)}")
        
        # If it's a range instruction, validate range
        if instruction.get('type') == 'range':
            start_q = instruction.get('start_question')
            end_q = instruction.get('end_question')
            
            if start_q is None or end_q is None:
                errors.append("Range instructions must have start_question and end_question")
            elif start_q > end_q:
                errors.append("start_question must be <= end_question")
        
        return len(errors) == 0, errors 